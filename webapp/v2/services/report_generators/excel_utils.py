"""
Excel导出工具类
提供通用的Excel文件生成功能，支持样式、格式化等
"""

import pandas as pd
import os
from typing import List, Dict, Any, Optional, Union
from datetime import datetime
from decimal import Decimal
import logging

logger = logging.getLogger(__name__)

class ExcelExportUtils:
    """Excel导出工具类"""
    
    @staticmethod
    def create_excel_file(
        data: List[Dict[str, Any]],
        file_path: str,
        sheet_name: str = "Sheet1",
        title: Optional[str] = None,
        subtitle: Optional[str] = None,
        columns_config: Optional[List[Dict[str, Any]]] = None,
        summary_data: Optional[Dict[str, Any]] = None,
        include_charts: bool = False
    ) -> str:
        """
        创建Excel文件
        
        Args:
            data: 数据列表
            file_path: 文件路径
            sheet_name: 工作表名称
            title: 报表标题
            subtitle: 报表副标题
            columns_config: 列配置信息
            summary_data: 汇总数据
            include_charts: 是否包含图表
            
        Returns:
            生成的文件路径
        """
        try:
            # 确保目录存在
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            
            if not data:
                # 创建空的Excel文件
                df = pd.DataFrame()
                df.to_excel(file_path, sheet_name=sheet_name, index=False)
                logger.warning(f"创建了空的Excel文件: {file_path}")
                return file_path
            
            # 转换数据为DataFrame
            df = pd.DataFrame(data)
            
            # 格式化数值列
            df = ExcelExportUtils._format_numeric_columns(df, columns_config)
            
            # 使用ExcelWriter进行高级格式化
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # 写入数据
                start_row = 0
                
                # 添加标题
                if title:
                    title_df = pd.DataFrame([[title]])
                    title_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, 
                                     index=False, header=False)
                    start_row += 2
                
                # 添加副标题
                if subtitle:
                    subtitle_df = pd.DataFrame([[subtitle]])
                    subtitle_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row,
                                       index=False, header=False)
                    start_row += 2
                
                # 写入主数据
                df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
                
                # 添加汇总数据
                if summary_data:
                    summary_start_row = start_row + len(df) + 3
                    ExcelExportUtils._add_summary_section(
                        writer, sheet_name, summary_data, summary_start_row
                    )
                
                # 应用样式
                ExcelExportUtils._apply_excel_styles(
                    writer, sheet_name, df, start_row, title, subtitle, columns_config
                )
                
                # 添加图表（如果需要）
                if include_charts and len(df) > 0:
                    ExcelExportUtils._add_charts(writer, sheet_name, df, start_row)
            
            logger.info(f"成功创建Excel文件: {file_path}, 数据行数: {len(df)}")
            return file_path
            
        except Exception as e:
            logger.error(f"创建Excel文件失败: {str(e)}")
            raise
    
    @staticmethod
    def _format_numeric_columns(
        df: pd.DataFrame, 
        columns_config: Optional[List[Dict[str, Any]]]
    ) -> pd.DataFrame:
        """格式化数值列"""
        if not columns_config:
            return df
        
        for col_config in columns_config:
            col_key = col_config.get('key')
            col_type = col_config.get('type')
            
            if col_key in df.columns:
                if col_type == 'currency':
                    # 格式化货币字段
                    df[col_key] = pd.to_numeric(df[col_key], errors='coerce').fillna(0)
                    df[col_key] = df[col_key].apply(lambda x: f"{x:.2f}")
                elif col_type == 'number':
                    # 格式化数字字段
                    df[col_key] = pd.to_numeric(df[col_key], errors='coerce').fillna(0)
                elif col_type == 'percentage':
                    # 格式化百分比字段
                    df[col_key] = pd.to_numeric(df[col_key], errors='coerce').fillna(0)
                    df[col_key] = df[col_key].apply(lambda x: f"{x:.2%}")
        
        return df
    
    @staticmethod
    def _add_summary_section(
        writer: pd.ExcelWriter,
        sheet_name: str,
        summary_data: Dict[str, Any],
        start_row: int
    ):
        """添加汇总部分"""
        summary_rows = []
        
        # 添加汇总标题
        summary_rows.append(['汇总信息', ''])
        summary_rows.append(['', ''])
        
        # 添加汇总数据
        for key, value in summary_data.items():
            if isinstance(value, (int, float, Decimal)):
                formatted_value = f"{value:.2f}" if isinstance(value, (float, Decimal)) else str(value)
            else:
                formatted_value = str(value)
            summary_rows.append([key, formatted_value])
        
        # 写入汇总数据
        summary_df = pd.DataFrame(summary_rows)
        summary_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row,
                           index=False, header=False)
    
    @staticmethod
    def _apply_excel_styles(
        writer: pd.ExcelWriter,
        sheet_name: str,
        df: pd.DataFrame,
        data_start_row: int,
        title: Optional[str],
        subtitle: Optional[str],
        columns_config: Optional[List[Dict[str, Any]]]
    ):
        """应用Excel样式"""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            workbook = writer.book
            worksheet = workbook[sheet_name]
            
            # 定义样式
            title_font = Font(name='微软雅黑', size=16, bold=True)
            subtitle_font = Font(name='微软雅黑', size=12, bold=True)
            header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
            data_font = Font(name='微软雅黑', size=10)
            
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            
            center_alignment = Alignment(horizontal='center', vertical='center')
            left_alignment = Alignment(horizontal='left', vertical='center')
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            current_row = 1
            
            # 应用标题样式
            if title:
                cell = worksheet.cell(row=current_row, column=1)
                cell.font = title_font
                cell.alignment = center_alignment
                # 合并标题行
                worksheet.merge_cells(f'A{current_row}:{get_column_letter(len(df.columns))}{current_row}')
                current_row += 2
            
            # 应用副标题样式
            if subtitle:
                cell = worksheet.cell(row=current_row, column=1)
                cell.font = subtitle_font
                cell.alignment = center_alignment
                # 合并副标题行
                worksheet.merge_cells(f'A{current_row}:{get_column_letter(len(df.columns))}{current_row}')
                current_row += 2
            
            # 应用表头样式
            header_row = data_start_row + 1
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=header_row, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            
            # 应用数据行样式
            for row_num in range(header_row + 1, header_row + len(df) + 1):
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.font = data_font
                    cell.border = thin_border
                    
                    # 根据列类型设置对齐方式
                    col_name = df.columns[col_num - 1]
                    if columns_config:
                        col_config = next((c for c in columns_config if c.get('key') == col_name), None)
                        if col_config and col_config.get('type') in ['currency', 'number', 'percentage']:
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                        else:
                            cell.alignment = left_alignment
                    else:
                        cell.alignment = left_alignment
            
            # 自动调整列宽
            for col_num in range(1, len(df.columns) + 1):
                column_letter = get_column_letter(col_num)
                max_length = 0
                
                # 检查列标题长度
                header_length = len(str(df.columns[col_num - 1]))
                max_length = max(max_length, header_length)
                
                # 检查数据长度
                for row_num in range(header_row + 1, header_row + min(len(df), 100) + 1):
                    cell_value = worksheet.cell(row=row_num, column=col_num).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                
                # 设置列宽（最小8，最大50）
                adjusted_width = min(max(max_length + 2, 8), 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
        except ImportError:
            logger.warning("openpyxl不可用，跳过样式设置")
        except Exception as e:
            logger.warning(f"应用Excel样式失败: {str(e)}")
    
    @staticmethod
    def _add_charts(
        writer: pd.ExcelWriter,
        sheet_name: str,
        df: pd.DataFrame,
        data_start_row: int
    ):
        """添加图表"""
        try:
            from openpyxl.chart import ColumnChart, Reference
            
            workbook = writer.book
            worksheet = workbook[sheet_name]
            
            # 查找数值列
            numeric_columns = []
            for col_name in df.columns:
                if df[col_name].dtype in ['int64', 'float64']:
                    numeric_columns.append(col_name)
            
            if len(numeric_columns) > 0:
                # 创建柱状图
                chart = ColumnChart()
                chart.title = "数据图表"
                chart.style = 13
                
                # 设置数据范围
                data_range = Reference(
                    worksheet,
                    min_col=2,  # 假设第一列是标签
                    min_row=data_start_row + 1,
                    max_col=min(len(numeric_columns) + 1, 5),  # 最多显示4个数据系列
                    max_row=data_start_row + min(len(df), 20) + 1  # 最多显示20行数据
                )
                
                categories = Reference(
                    worksheet,
                    min_col=1,
                    min_row=data_start_row + 2,
                    max_row=data_start_row + min(len(df), 20) + 1
                )
                
                chart.add_data(data_range, titles_from_data=True)
                chart.set_categories(categories)
                
                # 设置图表位置
                chart_position = f"{chr(ord('A') + len(df.columns) + 2)}{data_start_row + 1}"
                worksheet.add_chart(chart, chart_position)
                
        except ImportError:
            logger.warning("openpyxl图表功能不可用，跳过图表创建")
        except Exception as e:
            logger.warning(f"添加图表失败: {str(e)}")
    
    @staticmethod
    def format_currency(value: Union[int, float, Decimal, str]) -> str:
        """格式化货币值"""
        try:
            if isinstance(value, str):
                value = float(value)
            return f"{value:.2f}"
        except (ValueError, TypeError):
            return "0.00"
    
    @staticmethod
    def format_percentage(value: Union[int, float, Decimal, str]) -> str:
        """格式化百分比值"""
        try:
            if isinstance(value, str):
                value = float(value)
            return f"{value:.2%}"
        except (ValueError, TypeError):
            return "0.00%" 